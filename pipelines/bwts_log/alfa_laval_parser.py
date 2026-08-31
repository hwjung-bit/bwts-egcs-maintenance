# alfa_laval_parser.py — Alfa Laval PureBallast log parser
# Supports: XLSX (Tabular Export), A_*.csv (Alarm), E_*.csv (Event)
import re
import csv
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


# PureBallast Event Codes
EVENT_CODES = {
    # Operations
    "E10": "Operator command: Start Ballast",
    "E20": "Operator command: Start Deballast",
    "E190": "Standby mode",
    "E210": "Ballast Start-up mode",
    "E220": "Full Ballast mode",
    "E230": "Deballast mode",
    "E240": "Filter backflush",
    "E310": "E-stop",
    "E320": "Power on",
    "E370": "Request: Open overboard valve",
    "E380": "Request: Close overboard valve",
    "E390": "Request: Start ballast pump",
    "E400": "Request: Stop ballast pump",
    "E430": "Power used for process",
    "E450": "Shut down due to alarm",
    "E805": "System started in IMO mode",
    "E808": "Salinity changed",
}

# Ballast operation events
BALLAST_START = {"E10", "E210", "E220"}
DEBALLAST_START = {"E20", "E230"}
STOP_EVENTS = {"E190", "E310", "E450"}


def _read_xlsx_events(filepath):
    """PureBallast Tabular Export XLSX → (events, vessel_name, serial)."""
    if not openpyxl:
        return [], None, None

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row (Log time, GPS, Alarm/Event, Description)
    header_row = None
    vessel_name = None
    serial = None

    events = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Sheets exported by other tools have short/empty rows — pad so the
        # positional reads below never index past the tuple.
        row = tuple(row) + (None,) * max(0, 7 - len(row))
        if i == 3 and row[1]:
            vessel_name = str(row[1]).strip()
        if i == 3 and row[3]:
            serial = str(row[3]).strip()

        # Find header
        if row[0] and str(row[0]).strip() == "Log time":
            header_row = i
            continue

        if header_row is None:
            continue

        # Data rows
        ts = row[0]
        if not ts:
            continue
        if isinstance(ts, datetime):
            dt = ts
        else:
            try:
                dt = datetime.strptime(str(ts).strip(),
                                       "%Y-%m-%d %H:%M:%S")
            except (ValueError, TypeError):
                continue

        gps = str(row[1] or "").strip()
        code = str(row[2] or "").strip()
        desc = str(row[3] or "").strip()
        value = row[5] if len(row) > 5 else None
        unit = str(row[6] or "").strip() if len(row) > 6 else ""

        events.append({
            "time": dt,
            "date": dt.strftime("%Y-%m-%d"),
            "gps": gps if gps != "-" else None,
            "code": code,
            "description": desc,
            "value": value,
            "unit": unit,
            "is_alarm": code.startswith("A"),
        })

    wb.close()
    return events, vessel_name, serial


def parse_xlsx(filepath):
    """Parse one XLSX and analyze it (kept for callers/tests)."""
    events, vessel_name, serial = _read_xlsx_events(filepath)
    if not events and vessel_name is None:
        return None
    return _analyze_events(events, vessel_name, serial)


def parse_csv_alarm(filepath):
    """Parse A_*.csv (Alarm log, semicolon-separated)."""
    rows = _read_semicolon_csv(filepath)
    if len(rows) < 6:
        return []

    alarms = []
    for r in rows[5:]:  # Skip 4 header lines + column header
        if len(r) < 2:
            continue
        ts = _parse_ts(r[0])
        if not ts:
            continue
        alarms.append({
            "time": ts,
            "date": ts.strftime("%Y-%m-%d"),
            "alarm_number": r[1].strip() if len(r) > 1 else "",
            "is_alarm": True,
        })
    return alarms


def parse_csv_event(filepath):
    """Parse E_*.csv or IE_*.csv (Event log, semicolon)."""
    rows = _read_semicolon_csv(filepath)
    if len(rows) < 6:
        return []

    events = []
    for r in rows[5:]:
        if len(r) < 3:
            continue
        ts = _parse_ts(r[0])
        if not ts:
            continue
        events.append({
            "time": ts,
            "date": ts.strftime("%Y-%m-%d"),
            "gps": r[1].strip() if len(r) > 1 else None,
            "event_number": r[2].strip() if len(r) > 2 else "",
            "data": r[3].strip() if len(r) > 3 else "",
        })
    return events


def _read_semicolon_csv(filepath):
    """Read semicolon-separated CSV."""
    path = Path(filepath)
    for enc in ("utf-8-sig", "utf-8", "cp949"):
        try:
            with open(path, "r", encoding=enc) as f:
                return list(csv.reader(f, delimiter=";"))
        except UnicodeDecodeError:
            continue
    return []


def _parse_ts(s):
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _analyze_events(events, vessel_name=None, serial=None):
    """
    Analyze PureBallast events to extract operations.
    """
    if not events:
        return {
            "vessel_name": vessel_name,
            "serial": serial,
            "ballast_count": 0,
            "deballast_count": 0,
            "op_days": 0,
            "alarms": [],
            "alarm_count": 0,
            "operations": [],
            "issues": [],
        }

    # Count operations
    operations = []
    current_op = None
    ballast_dates = set()
    deballast_dates = set()

    for ev in sorted(events, key=lambda e: e["time"]):
        code = ev.get("code", "")

        if code in BALLAST_START:
            if current_op and current_op["mode"] == "BALLAST":
                continue  # Already in ballast
            if current_op:
                current_op["end_time"] = ev["time"]
                operations.append(current_op)
            current_op = {
                "mode": "BALLAST",
                "start_time": ev["time"],
                "date": ev["date"],
                "end_time": None,
            }
            ballast_dates.add(ev["date"])

        elif code in DEBALLAST_START:
            if current_op and current_op["mode"] == "DEBALLAST":
                continue
            if current_op:
                current_op["end_time"] = ev["time"]
                operations.append(current_op)
            current_op = {
                "mode": "DEBALLAST",
                "start_time": ev["time"],
                "date": ev["date"],
                "end_time": None,
            }
            deballast_dates.add(ev["date"])

        elif code in STOP_EVENTS:
            if current_op:
                current_op["end_time"] = ev["time"]
                operations.append(current_op)
                current_op = None

    if current_op:
        operations.append(current_op)

    # Alarms
    alarms = [ev for ev in events if ev.get("is_alarm")]
    alarm_codes = {}
    for a in alarms:
        c = a.get("code") or a.get("alarm_number", "")
        alarm_codes[c] = alarm_codes.get(c, 0) + 1

    # Issues
    issues = []
    estops = sum(1 for ev in events
                 if ev.get("code") == "E310")
    shutdowns = sum(1 for ev in events
                    if ev.get("code") == "E450")
    if estops > 3:
        issues.append(f"E-stop {estops}회")
    if shutdowns > 0:
        issues.append(f"알람 Shutdown {shutdowns}회")

    # Salinity info
    salinity_events = [
        ev for ev in events if ev.get("code") == "E808"]
    salinities = []
    for se in salinity_events:
        v = se.get("value")
        if v and isinstance(v, (int, float)):
            salinities.append(float(v))

    return {
        "vessel_name": vessel_name,
        "serial": serial,
        "ballast_count": len([o for o in operations
                              if o["mode"] == "BALLAST"]),
        "deballast_count": len([o for o in operations
                                if o["mode"] == "DEBALLAST"]),
        "op_days": len(ballast_dates | deballast_dates),
        "operations": operations,
        "alarms": alarms,
        "alarm_count": len(alarms),
        "alarm_codes": alarm_codes,
        "estop_count": estops,
        "shutdown_count": shutdowns,
        "issues": issues,
        "salinity_avg": (
            round(sum(salinities) / len(salinities), 1)
            if salinities else None),
    }


# Periodic measurement events written once a minute while the system runs.
# E640 = treated flow [m3/h]; when it is > 0 the plant was operating even if
# the export window missed the E10/E210/E220 start event.
FLOW_EVENT = "E640"
# Housekeeping codes that appear in an idle month too
HEARTBEAT_CODES = {"E104", "E190", "E320"}


def _in_month(ev, year, month):
    t = ev.get("time")
    return year is None or (t and t.year == year and t.month == month)


def analyze_alfa_laval(folder_path, year=None, month=None):
    """Analyze an Alfa Laval vessel folder.

    Reads EVERY source in the folder (all XLSX exports and all A_/E_/IE_
    CSVs), merges them, drops duplicates and — when year/month are given —
    keeps only that month's events. The old "first XLSX wins" rule lost
    September 2025 for KDB, whose folder also held a multi-year CSV dump,
    and a month-spanning CSV would have been counted for the wrong month.

    Extra keys on the result:
      event_count      events in the month (alarms excluded)
      flow_rows        E640 rows with flow > 0 (operating evidence)
      sources          files used
    """
    folder = Path(folder_path)
    if not folder.exists():
        return None

    events, sources = [], []
    vessel_name = serial = None
    for f in sorted(folder.glob("*.xlsx")):
        if f.name.startswith("~$"):
            continue
        try:
            evs, vn, sn = _read_xlsx_events(f)
        except Exception as e:      # a non-PureBallast workbook in the folder
            print(f"  [alfa] {f.name}: 읽기 실패 ({e}) — 건너뜀")
            continue
        if evs:
            events.extend(evs)
            sources.append(f.name)
        vessel_name = vessel_name or vn
        serial = serial or sn
    for pattern in ("A_*.csv", "E_*.csv", "IE_*.csv"):
        for f in sorted(folder.glob(pattern)):
            if pattern.startswith("A"):
                rows = parse_csv_alarm(f)
                for a in rows:
                    a["code"] = f"A{a.get('alarm_number', '')}"
            else:
                rows = parse_csv_event(f)
                for e in rows:
                    e["code"] = f"E{e.get('event_number', '')}"
            if rows:
                events.extend(rows)
                sources.append(f.name)

    if not events:
        return None

    # de-duplicate (same event exported twice) and month-filter
    seen, merged = set(), []
    for ev in events:
        key = (ev.get("time"), ev.get("code"), str(ev.get("data", ev.get("value", ""))))
        if key in seen:
            continue
        seen.add(key)
        if _in_month(ev, year, month):
            merged.append(ev)

    result = _analyze_events(merged, vessel_name, serial)
    flow_rows = 0
    for ev in merged:
        if ev.get("code") == FLOW_EVENT:
            try:
                if float(str(ev.get("data", ev.get("value", 0)) or 0)) > 0:
                    flow_rows += 1
            except ValueError:
                pass
    result["flow_rows"] = flow_rows
    result["event_count"] = sum(
        1 for ev in merged if not ev.get("is_alarm") and ev.get("code") not in HEARTBEAT_CODES)
    # E104 is logged every few days while the system is powered: several of
    # them spread over the month prove the export covered the month.
    hb_days = {ev["time"].date() for ev in merged if ev.get("code") == "E104" and ev.get("time")}
    result["heartbeat_days"] = len(hb_days)
    result["sources"] = sources
    result["source"] = ", ".join(sources[:3]) + (" …" if len(sources) > 3 else "")
    return result
